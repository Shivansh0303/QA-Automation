import openpyxl
import os
import time
import tkinter as t
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver 
from selenium.webdriver.chrome.service import Service as ChromeService

def func(r,email1,password1,directory = os.path.dirname(os.path.realpath(__file__))) :
    
    links = []
    mails = []

    db_object = openpyxl.load_workbook(directory + "\\" + r)
    mail_object = db_object.active
    mail_rows = mail_object.max_row
    for n in range(2,mail_rows+1) :
        if ((mail_object.cell(column=1 , row=n)).value) != None :
            mails.append((mail_object.cell(column=1 , row=n)).value)

    names = []
    place = []
    experience = []

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.linkedin.com/login")
    driver.find_element(By.ID, "username").send_keys(email1)
    driver.find_element(By.ID, "password").send_keys(password1)
    driver.find_element(By.XPATH, "//*[@id='organic-div']/form/div[3]/button").click()
    time.sleep(29)
    driver.implicitly_wait(5)
    for mail in mails:
        driver.get(mail)
        driver.implicitly_wait(7)
        time.sleep(5)
        names.append(driver.find_element(By.XPATH,'//h1').get_attribute("innerHTML"))
        place.append(driver.find_element(By.XPATH,'//div[@class="pv-text-details__left-panel mt2"]/span[1]').text)
        sections_elements = driver.find_elements(By.XPATH,'//section[@data-view-name="profile-card"]/div[@class="pvs-header__container"]/div/div/div/h2/span[1]')
        section_name = []
        #/div/ul/li)[1]/div/div/div
        for section_element in sections_elements :
            presence = [0,1]
            section_name.append(section_element.text)
            if section_name[-1] == 'Experience':
                presence.append(1)
                if len(driver.find_elements(By.XPATH,f"(/html//section[@data-view-name='profile-card'][{len(section_name)}]/div/ul/li)[1]/div/div/div")) == 1 :
                    experience.append(driver.find_element(By.XPATH,f"(//section[@data-view-name='profile-card'][{len(section_name)}]/div/ul/li)[1]/div/div/div").text)

                elif len(driver.find_elements(By.XPATH,f'(//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div/div/div')) == 2 :

                    if len(driver.find_elements(By.XPATH, f'((//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div/div/div)[2]/ul/li')) == 1 :
                        experience.append(driver.find_element(By.XPATH,f'((//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div/div/div)[1]').text)
                    elif len(driver.find_elements(By.XPATH, f'((//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div/div/div)[2]/ul/li')) >= 2:
                        experience.append(driver.find_element(By.XPATH,f'((//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div/div/div//span[@aria-hidden="true"])[1]').text)
                        experience[-1] += " "+(driver.find_element(By.XPATH,f'((//section[@data-view-name="profile-card"][{len(section_name)}]/div/ul/li)[1]/div//div[@class="pvs-list__outer-container"]/ul/li)[1]').text)
            if section_element == sections_elements[-1] and len(presence) == 3:
                experience.append('N/A')

        print(section_name)
        print(experience)
    
    print(names)
    print(place)
    print(len(names))
    print(len(place))
    print(len(experience))
    


    df = {'links':mails,'name':names,'place':place,'experience':experience}
    df = pd.DataFrame(df)
    df.to_excel(f"{r} output.xlsx")

top = t.Tk()
L1 = t.Label(top, text = "Excel File").grid(row =0, column =0)
E1 = t.Entry(top, bd = 5)
E1.grid(row =0, column =1)
L4 = t.Label(top, text = "Email LinkedIn").grid(row =1, column =0)
E4 = t.Entry(top, bd = 7)
E4.grid(row =1, column =1)
L5 = t.Label(top, text = "Password").grid(row =2, column =0)
E5 = t.Entry(top,show = "*", bd = 7)
E5.grid(row=2, column =1)
t.Button(top, text= "Run", command = lambda : func(E1.get(),E4.get(),E5.get())).grid(row = 3, column =1)
top.mainloop()

top.mainloop()