import openpyxl
import os
import tkinter as t
import time
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager  
from selenium import webdriver 
from selenium.webdriver.chrome.service import Service as ChromeService


def func(r,email,password,email1,password1,directory = os.path.dirname(os.path.realpath(__file__))) :
    
    links = []
    mails = []

    db_object = openpyxl.load_workbook(directory + "\\" + r)
    mail_object = db_object.active
    mail_rows = mail_object.max_row
    for n in range(1,mail_rows+1) :
        if ((mail_object.cell(column=1 , row=n)).value) != None :
            mails.append((mail_object.cell(column=1 , row=n)).value)

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.get("https://outlook.live.com/owa/")
    driver.find_element(By.LINK_TEXT, "Sign in").click()
    driver.find_element(By.ID, "i0116").send_keys(email)
    driver.find_element(By.ID, "idSIButton9").click()
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//input[@name='passwd']").send_keys(password)
    driver.find_element(By.XPATH, "//input[@value = 'Sign in']").click()
    driver.find_element(By.ID, "idBtn_Back").click()
    
    driver.switch_to.new_window()
    driver.get("https://www.linkedin.com/login")
    driver.find_element(By.ID, "username").send_keys(email1)
    driver.find_element(By.ID, "password").send_keys(password1)
    driver.find_element(By.XPATH, "//*[@id='organic-div']/form/div[3]/button").click()
    driver.implicitly_wait(5)
    time.sleep(4)
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//div[@data-automation-type='RibbonSplitButton']//button/span").click()
    asd = 0
    name= []
    connections = []
    description = []
    images = []
    designations = []
    joining_date = []
    profile_update = []
    contact_update = []
    
    def linkedin_extract(driver) :
        driver.implicitly_wait(5)
        try:
            connections.append(driver.find_element(By.XPATH, '//span[@class="t-bold"]').get_attribute("innerHTML"))
        except:
            connections.append("0")
        '''
        driver.find_element(By.XPATH, '//div[@class="pv-top-card-v2-ctas "]//span[text()="More"]').click()
        driver.find_element(By.XPATH, '//main/section[1]//div/span[text()="About this profile"][1]').click()
        driver.implicitly_wait(5)

        if len(driver.find_elements(By.XPATH, "//ul[@class='list-style-none pt1']/li")) >= 3:
            joining_date.append(" ".join(driver.find_element(By.XPATH, "//ul[@class='list-style-none pt1']/li[1]/span/span[1]").text.split()[1::]))
            print(joining_date[-1])
            profile_update.append(" ".join(driver.find_element(By.XPATH, "//ul[@class='list-style-none pt1']/li[2]/span/span[1]").text.split()[2::]))
            print(profile_update[-1])
            contact_update.append(" ".join(driver.find_element(By.XPATH, "//ul[@class='list-style-none pt1']/li[3]/span/span[1]").text.split()[2::]))
            print(contact_update[-1])

        elif len(driver.find_elements(By.XPATH, "//ul[@class='list-style-none pt1']/li")) == 2:
            joining_date.append(" ".join(driver.find_element(By.XPATH, "//ul[@class='list-style-none pt1']/li[1]/span/span[1]").text.split()[1::]))
            print(joining_date[-1])
            profile_update.append(" ".join(driver.find_element(By.XPATH, "//ul[@class='list-style-none pt1']/li[2]/span/span[1]").text.split()[2::]))
            print(profile_update[-1])
            contact_update.append("N/A")
            print(contact_update[-1])
        '''

    for mail in mails :
        driver.find_element(By.XPATH,"//div[@class='UKx9j']//div[@role='textbox']").send_keys(mail + ";")
        asd+=1

        if asd == 1:
                time.sleep(5)
        
        while True:
            try:
                action = ActionChains(driver)
                action.move_to_element(driver.find_element(By.XPATH, f"//span[text()= '{mail}']")).perform()
                action.context_click().perform()
                driver.find_element(By.XPATH, "//span[text()='Open contact card']").click()
                driver.find_element(By.XPATH, '//button[@data-content="LinkedIn"]//span[text()="LinkedIn"]').click()
                break
            except:
                continue
        driver.implicitly_wait(5)
        #driver.find_element(By.XPATH, "//section[@aria-label='LinkedIn']//button").click()
        #if len(driver.find_elements(By.CSS_SELECTOR, "button > svg")) == 1:

        if len(driver.find_elements(By.XPATH, "//section//button//span")) == 3:
            links.append('N/A')
            name.append('N/A')
            images.append("N/A")
            description.append('N/A')
            designations.append('N/A')
            connections.append('N/A')
            #joining_date.append('N/A')
            #profile_update.append('N/A')
            #contact_update.append('N/A')
            
        elif len(driver.find_elements(By.XPATH, "//section//button//span")) == 7 and len(driver.find_elements(By.CSS_SELECTOR, "section[aria-label='LinkedIn'] button")) == 1:
            name.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div/div').get_attribute("innerHTML"))
            print(name[-1])
            try:
                images.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]//img').get_attribute("src"))
            except:
                images.append("N/A")
            description.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[2]').get_attribute("innerHTML"))
            designations.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[3]/div').get_attribute("innerHTML"))
            '''
            if driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[3]/div[3]').get_attribute("innerHTML") == "0 connections":
                time.sleep(1)
            '''
            driver.find_element(By.CSS_SELECTOR, "section[aria-label='LinkedIn'] button").click()
            driver.switch_to.window(driver.window_handles[1])
            links.append(driver.current_url)
            linkedin_extract(driver)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        else:
            try:
                time.sleep(1)
                element = driver.find_element(By.XPATH, "//*[@id='ImmersiveProfileOverlayWrapper']/div/div/div[2]/div/div[2]")
                driver.execute_script('arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;', element)
            except:
                None
            name.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div/div').get_attribute("innerHTML"))
            print(name[-1])
            try:
                images.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]//img').get_attribute("src"))
            except:
                images.append("N/A")
            description.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[2]').get_attribute("innerHTML"))
            designations.append(driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[3]/div').get_attribute("innerHTML"))
            '''
            if driver.find_element(By.XPATH, '//section[@data-log-name="LinkedInHeader"]/div/div/div[2]/div[3]/div[3]').get_attribute("innerHTML") == "0 connections":
                time.sleep(1)
            '''
            driver.find_element(By.CSS_SELECTOR, "button > svg").click()
            driver.switch_to.window(driver.window_handles[1])
            links.append(driver.current_url)
            linkedin_extract(driver)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

        driver.find_element(By.CSS_SELECTOR, "div button[data-log-name='CloseButton']").click()
        driver.find_element(By.XPATH,"//div[@class='UKx9j']//div[@role='textbox']").send_keys(Keys.DELETE)
    
    print(len(connections))
    print(len(joining_date))
    print(len(profile_update))
    print(len(joining_date))
    print(connections)
    print(joining_date)
    print(profile_update)
    print(joining_date)

    wb_ob = openpyxl.Workbook()
    sheet_ob = wb_ob.active

    c1 = sheet_ob.cell(row = 1, column = 1)
    c1.value = "Email"
    c2 = sheet_ob.cell(row = 1, column = 2)
    c2.value = "Link"
    c3 = sheet_ob.cell(row = 1, column = 3)
    c3.value = "Name"
    c4 = sheet_ob.cell(row = 1, column = 4)
    c4.value = "connections"
    c5 = sheet_ob.cell(row = 1, column = 5)
    c5.value = "Designation"
    c6 = sheet_ob.cell(row = 1, column = 6)
    c6.value = "Location"
    c7 = sheet_ob.cell(row = 1, column = 7)
    c7.value = "Image"
    c8 = sheet_ob.cell(row = 1, column = 8)
    c8.value = "joining date"
    c9 = sheet_ob.cell(row = 1, column = 9)
    c9.value = "Picture Update"
    c10 = sheet_ob.cell(row = 1, column = 10)
    c10.value = "Contact update"
    

    for num in range(len(mails)):
        (sheet_ob.cell(row = num+2, column =1)).value = mails[num]
        (sheet_ob.cell(row = num+2, column =2)).value = links[num]
        (sheet_ob.cell(row = num+2, column =3)).value = name[num]
        (sheet_ob.cell(row = num+2, column =4)).value = connections[num]
        (sheet_ob.cell(row = num+2, column =5)).value = description[num]
        (sheet_ob.cell(row = num+2, column =6)).value = designations[num]
        (sheet_ob.cell(row = num+2, column =7)).value = images[num]
        #(sheet_ob.cell(row = num+2, column =8)).value = joining_date[num]
        #(sheet_ob.cell(row = num+2, column =9)).value = contact_update[num]
        #(sheet_ob.cell(row = num+2, column =10)).value = profile_update[num]

    wb_ob.save(f"{r} - output.xlsx")

top = t.Tk()
top.title("Outlooker©")
top.geometry('250x200')
L1 = t.Label(top, text = "Excel File").grid(row =0, column =0)
E1 = t.Entry(top, bd = 5)
E1.grid(row =0, column =1)
L2 = t.Label(top, text = "Email Outlook").grid(row =1, column =0)
E2 = t.Entry(top, bd = 7)
E2.grid(row =1, column =1)
L3 = t.Label(top, text = "Password").grid(row =2, column =0)
E3 = t.Entry(top,show = "*", bd = 7)
E3.grid(row=2, column =1)
L4 = t.Label(top, text = "Email LinkedIn").grid(row =3, column =0)
E4 = t.Entry(top, bd = 7)
E4.grid(row =3, column =1)
L5 = t.Label(top, text = "Password").grid(row =4, column =0)
E5 = t.Entry(top,show = "*", bd = 7)
E5.grid(row=4, column =1)
t.Button(top, text= "Run", command = lambda : func(E1.get(),E2.get(),E3.get(),E4.get(),E5.get()) ).grid(row = 5, column =1)
top.mainloop()