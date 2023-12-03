import openpyxl
import re   
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

ignore_domains = ["gmail.com" , "yahoo.com", "me.com", "icloud.com","outlook.com"];
r = input("Enter the excel file name with extension : ")
wb_object = openpyxl.load_workbook(r);
sheet_object = wb_object.active;
chrome_driver_path = "/path/to/chromedriver"  # Update this path
rows = sheet_object.max_row;
columns = sheet_object.max_column;

s=[];
for n in range(2,rows+1) :
    if ((sheet_object.cell(column=2 , row=n)).value) != None :
        s.append((sheet_object.cell(column=2 , row=n)).value)

g=[];
for n in range(2,rows+1) :
    if ((sheet_object.cell(column=1 , row=n)).value) != None :
        g.append((sheet_object.cell(column=1,row=n)).value);

domain=[];
url_s=[];
for i in range(len(s)) :
    count=[];
    for j in range(len(ignore_domains)) :
            
        if ignore_domains[j] in s[i] :
            domain.append(s[i]);
            domain.append("Not a buisness email");
            
        else:
            count.append(1);
    r='';
    if len(count) == 5:
        match = re.search(r'@\S+',s[i]);
        r=match.group();
        domain.append(s[i]);
        domain.append(r[1::]);
        url_s.append(r[1::]);

url_s.append("");

workbook = Workbook();
workbook.save(filename="sample.xlsx");

options = webdriver.ChromeOptions();
options.add_experimental_option("detach", True);
driver = webdriver.Chrome(executable_path=chrome_driver_path, options=options)
results = [];
dom_results = [];

a = 0;
for b in range(len(domain)) :

    if domain[b] != "Not a buisness email" :
        driver.get("http://google.com/");
        driver.find_element_by_name("q").send_keys(f'"{domain[b]}"',Keys.ENTER);
        text = driver.page_source;
        occurences = re.findall(f"<em>{domain[b]}</em>",text);

        if len(occurences) != 0 :
            dta = driver.find_elements_by_xpath("//div[@class='yuRUbf']//a[not(@class)]");

            links=[];
            for output in dta:
                links.append(output.get_attribute("href"));

            if url_s[a] == domain[b] :
                dom_results.append(links);
                a+=1
            else:
                results.append(links);

        else :
                
            if url_s[a] == domain[b] :
                    dom_results.append(["N/A"]);
                    a+=1
            else:
                    results.append(["N/A"]);
        
    else: 
        dom_results.append(["N/A"]);

a = 0;

wb_ob = openpyxl.Workbook() ;
sheet_ob = wb_ob.active;

c1 = sheet_ob.cell(row = 1, column =1);
c1.value = "Name";
c2 = sheet_ob.cell(row = 1, column =2);
c2.value = "Email";
c3 = sheet_ob.cell(row = 1, column =3);
c3.value = "Exact searches";
c4 = sheet_ob.cell(row = 1, column =4);
c4.value = "Domain searches";

z1 = 2;
z2 = 2;
for num in range(len(s)) :
    (sheet_ob.cell(row = z1, column =1)).value = g[num];
    (sheet_ob.cell(row = z1, column =2)).value = s[num];

    for digit in range(len(results[num])) :
        (sheet_ob.cell(row = z1, column =3)).value = results[num][digit];
        z1+=1;
        
    for digit in range(len(dom_results[num])) :
        (sheet_ob.cell(row = z2, column =4)).value = dom_results[num][digit];
        z2+=1;

    if z1 > z2 :
        z2 = z1;
    else:
        z1 = z2;

driver.quit();
wb_ob.save("output.xlsx");